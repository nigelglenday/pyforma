"""Tie-out: every modeled line must match the source workbook, every year.

This is the test that matters. The Python model is only worth anything if it
reproduces `reference/Standard Operating Model v2.xlsx` exactly, so this reads
the workbook's own cached values and compares them line by line.

Run with: python3 -m pytest tests/ -v      (or: python3 tests/test_tie_out.py)
"""

import openpyxl
import pytest
import yaml

from models.model import run_all

WORKBOOK = "reference/Standard Operating Model v2.xlsx"
CONFIG = "config/assumptions.yaml"

FIRST_COL = 6   # column F, the first projected year (2013)
YEARS = 10

# source row -> attribute path on the solved model
LINES = {
    90: ("balance_sheet", "cash"),
    91: ("balance_sheet", "securities"),
    92: ("balance_sheet", "loans_gross"),
    93: ("balance_sheet", "loan_loss_reserve"),
    94: ("balance_sheet", "loans_net"),
    98: ("balance_sheet", "fixed_assets"),
    99: ("balance_sheet", "other_assets"),
    100: ("balance_sheet", "total_assets"),
    103: ("balance_sheet", "deposits"),
    104: ("balance_sheet", "borrowings"),
    105: ("balance_sheet", "other_liabilities"),
    106: ("balance_sheet", "total_liabilities"),
    110: ("balance_sheet", "common_equity"),
    49: ("funding", "deposit_interest_expense"),
    54: ("funding", "borrowing_interest_expense"),
    70: ("income", "interest_income_on_loans"),
    71: ("income", "interest_income_on_securities_and_cash"),
    72: ("income", "interest_income"),
    73: ("income", "interest_expense"),
    74: ("income", "net_interest_income_before_provisions"),
    75: ("income", "loan_loss_provision"),
    76: ("income", "net_interest_income_after_provisions"),
    77: ("income", "non_interest_income"),
    78: ("income", "total_revenue"),
    79: ("income", "operating_expense"),
    82: ("income", "pretax_income"),
    83: ("income", "tax_provision"),
    84: ("income", "net_income"),
}

# Balances are in millions. A nanodollar is far below the model's precision,
# and well above the float noise of a 10-year compounding chain.
TOLERANCE = 1e-9


@pytest.fixture(scope="module")
def solved():
    with open(CONFIG) as f:
        return run_all(yaml.safe_load(f))


@pytest.fixture(scope="module")
def source():
    return openpyxl.load_workbook(WORKBOOK, data_only=True)["model"]


@pytest.mark.parametrize("row,path", sorted(LINES.items()))
def test_line_ties_to_source(solved, source, row, path):
    tab, field = path
    modeled = getattr(getattr(solved, tab), field)
    label = source.cell(row, 2).value
    for year in range(YEARS):
        expected = source.cell(row, FIRST_COL + year).value
        assert expected is not None, f"source row {row} year {year + 1} is empty"
        actual = modeled.iloc[year]
        assert abs(actual - expected) < TOLERANCE, (
            f"{label!r} (row {row}) year {year + 1}: "
            f"source {expected:.6f}, model {actual:.6f}, diff {actual - expected:.2e}"
        )


def test_balance_sheet_balances(solved):
    bs = solved.balance_sheet
    check = (bs.total_liabilities + bs.common_equity - bs.total_assets).abs().max()
    assert check < TOLERANCE, f"balance sheet off by {check:.2e}"


def test_source_check_cell_is_zero(source):
    """The workbook's own check cell (row 113) must be zero, or our target is wrong."""
    for year in range(YEARS):
        assert abs(source.cell(113, FIRST_COL + year).value) < TOLERANCE


def test_borrowings_plug_converged(solved):
    assert solved.iterations < 100, "borrowings plug hit the iteration cap"


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
